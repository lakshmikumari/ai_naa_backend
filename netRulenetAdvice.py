import os
import re
import openpyxl
from log import logger
import DB
import work_files
import json

# script path
scriptPath = os.path.dirname(os.path.realpath(__file__))


# function to write into a file
def writeData(file_name, data):
    with open(file_name, "w") as f:
        f.write(json.dumps(data))


def netRule_netAdvice_json(audit_info):
    """
    this function will fetch the net rule and net device from 'Audit_Summary_Excel.xlsx' under "Audit_Exception" tab
    output will be stored in netRule_netRule collection in DB
    """
    try:
        #netrule_data = DB.get_data('netRule_netAdvice')
        netrule_data=[]
        file_name = audit_info['fileName'].split('.')[0]
        jsonData = {}
        jsonDataList = []
        with open(work_files.dataDir + "/nms_map.json", "r") as f:
            nmsdata = json.loads(f.read())
        entries_to_remove = ('auditId', 'crPartyId', 'crPartyName', 'location')
        for k in entries_to_remove:
            nmsdata.pop(k)

        for audit_name in nmsdata:
            to_add = []
            for inner_key in nmsdata[audit_name].keys():
                to_add += nmsdata[audit_name][inner_key]
                # to_add.append(nms[key][inner_key])
            jsonData[audit_name] = to_add
        '''
        for x in nmsdata.keys():
            for i in nmsdata[x]:
                for j in nmsdata[x][i]:
                    list_data.append(j)
            if x not in jsonData.keys():
                jsonData[x] = list_data
            else:
                jsonData[x].append(list_data)
        '''
        wb = openpyxl.load_workbook(work_files.auditSummaryDir + 'Audit_Summary_Excel.xlsx')
        active_sheet = wb["Audit_Exception"]
        for i in range(5, active_sheet.max_row + 1):

            exception = active_sheet.cell(row=i, column=1).value
            exception = exception[exception.rfind(',"') + 2:-2]
            if "<br/>" in exception:
                exception = exception.replace("<br/>", "")
            table_name = active_sheet.cell(row=i, column=2).value
            table_name = re.sub(' +', ' ', table_name)
            net_rule = active_sheet.cell(row=i, column=7).value
            net_advice = active_sheet.cell(row=i, column=8).value
            audit_name = 'unknown'
            audit_names = []
            tempjson_list = []
            for k in jsonData.keys():
                if table_name in jsonData[k]:
                    # audit_name = k
                    audit_names.append(k)
                    # break

            if len(audit_names) == 0:
                tempJson = {"exception": exception, "table_name": table_name, "audit_name": audit_name,
                            "net_rule": net_rule,
                            "net_advice": net_advice}
                tempjson_list.append(tempJson)
                logger.info('')
            else:
                for audit in audit_names:
                    tempJson = {"exception": exception, "table_name": table_name, "audit_name": audit,
                                "net_rule": net_rule,
                                "net_advice": net_advice}
                    tempjson_list.append(tempJson)
            audits = [i['audit_name'] for i in netrule_data]
            test_list = {'audit_name': audit_name, 'table_name': table_name}
            if len(netrule_data) > 0:
                if audit_name not in audits:
                    jsonDataList.extend(tempjson_list)
                else:
                    if test_list not in netrule_data:
                        jsonDataList.extend(tempjson_list)
            else:
                jsonDataList.extend(tempjson_list)

        jsonDataList = [dict(t) for t in {tuple(d.items()) for d in jsonDataList}]
        if len(jsonDataList) > 0:
            writeData(work_files.netRulenetAdvice(scriptPath, file_name), jsonDataList)
            file_size = os.stat(work_files.netRulenetAdvice(scriptPath, file_name)).st_size
            logger.info("netRule-netInventory.json craeted, size: {}".format(round(file_size / 1024, 2)))
    except Exception as e:
        logger.info(e)

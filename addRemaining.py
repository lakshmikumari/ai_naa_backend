import os
from bs4 import BeautifulSoup
import work_files
import json
import openpyxl
from log import logger
from concurrent.futures import ThreadPoolExecutor
from concurrent.futures import ProcessPoolExecutor
import concurrent.futures

# script path
scriptPath = os.path.dirname(os.path.realpath(__file__))


# function to write into a file
def writeData(file_name, data):
    with open(file_name, "w") as f:
        f.write(json.dumps(data))


def processNodeFolders(tablename, folders, file_name, audit_info):
    for folder in folders:
        matchingFile = getMatchingFileName(folder, file_name, tablename)
        logger.info(f"{tablename} present in {matchingFile} in {folder}")

        for f in os.listdir(scriptPath + "/" + file_name + "/Full/" + folder):
            if f == matchingFile:
                logger.info(f"processing File : {f} folder : {folder}")
                with open(scriptPath + "/" + file_name + "/Full/" + folder + "/" + f, 'r') as F:
                    html_data = F.read()
                    soup = BeautifulSoup(html_data, features="html.parser")
                    tables = soup.find_all("table")

                    for table in tables:

                        if tablename in table.getText():
                            logger.info(f"tablename  in file  text: {f}")
                            # if the table name has / in it, it has to be removed as the filename to be created
                            # cannot have a /
                            tablename1 = tablename.replace("/", "")
                            # tableData = str(table.encode("utf-8")).replace("/xa0", " ")
                            if "Detailed_Findings" in f:
                                audit_name = f.split('_Detailed_Findings')[0]
                            elif "NodeAndSummary_" in f:
                                audit_name = (f.split('NodeAndSummary_')[1]).split('.')[0]
                            logger.info(f"audit_name: {audit_name} tablename : {tablename}")
                            filename_to_save = folder + "_" + tablename1 + "-" + audit_name
                            with open(work_files.addRemainingDir + filename_to_save + '.html', 'w') as w:
                                # tableData = str(table.encode("utf-8")).replace("/xa0", " ")
                                tableData = str(table)
                                w.write(tableData)
                            # logger.info('-- all html file has been created for add remaining---')
                            break

                break


def CreatefilesParallel(tdict, file_name, audit_info):
    # file_name = audit_info['fileName'].split('.')[0]
    os.chdir(work_files.exceptionsDir)
    try:
        with ThreadPoolExecutor(max_workers=3) as exe:
            results = [exe.submit(processNodeFolders, tablename, folders, file_name, audit_info)
                       for tablename, folders in tdict.items()]

    except Exception as e:
        logger.error(e)


def getMatchingFileName(folder, file_name, tablename):
    """
    function to check in which node based html, the table name is present
    Expected that the table may be present only in 1 html file under that device folder
    """
    for htmlfile in os.listdir(scriptPath + "/" + file_name + "/Full/" + folder):
        if not htmlfile.endswith(".html"):
            continue
        with open(scriptPath + "/" + file_name + "/Full/" + folder + "/" + htmlfile, 'r') as F:
            html_data = F.read()
            soup = BeautifulSoup(html_data, features="html.parser")
            tables = soup.find_all("table")
            for table in tables:
                head = table.find_all("thead")
                for h in head:
                    rows = h.findChildren('tr')
                    tableNames = [x.getText() for x in rows[0].findChildren('td')]
                    if tablename in tableNames:
                        logger.info(f"{tablename} present in {htmlfile} ")
                        return htmlfile


def add_remaining_files(audit_info):
    """input:  reads .html files from fileName=153875 + Full folder and sub folders
    Writes node name to model linking data like {"node name": "N9K-1", "Model": "cevChassisN9Kc9504"} to DB """
    try:
        table_dict = {}
        table_list = []
        file_name = audit_info['fileName'].split('.')[0]
        with open(work_files.internalDir + "/allData.json", "r") as f:
            allData_exception_json = json.loads(f.read())
        for json_data in allData_exception_json:
            table_list.append(json_data['tableName'])
        wb = openpyxl.load_workbook(work_files.auditSummaryDir + 'Audit_Summary_Excel.xlsx')
        active_sheet = wb["Audit_Exception"]
        for i in range(5, active_sheet.max_row + 1):
            table_name = active_sheet.cell(row=i, column=2).value
            if table_name not in table_list:
                logger.info('%s not present in summary folders', table_name)
                nodes_sheet = wb[f'Exception Row-{i} nodes']
                for j in range(4, nodes_sheet.max_row + 1):
                    device = nodes_sheet.cell(row=j, column=1).value
                    # logger.info(device)
                    device_dir = device.split('), "')
                    # logger.info(device_dir)
                    # deviceName = device_dir[1].split(',')[1][1:-2]
                    deviceFolder = device_dir[1].split('/')[0]

                    if table_name in table_dict:
                        table_dict[table_name].append(deviceFolder)
                    else:
                        table_dict[table_name] = [deviceFolder]
            else:
                pass

        CreatefilesParallel(table_dict, file_name, audit_info)

    except Exception as e:
        logger.error("**** Error while generating addremaining.json file ****".format(e))
